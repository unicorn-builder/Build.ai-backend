"""
gen_boq_xlsx.py — BOQ Structure as Excel (.xlsx) — Bilingual FR/EN
Tijan AI — same data as gen_boq_structure.py, Excel output via openpyxl
"""
import io, math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers


VERT_FILL = PatternFill(start_color="43A956", end_color="43A956", fill_type="solid")
GRIS_FILL = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=10, bold=True)
NORMAL_FONT = Font(name="Calibri", size=10)
NUM_FMT = '#,##0'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

# Translation helpers
def _t(fr, en, lang='fr'):
    return en if lang == 'en' else fr


def _add_header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = VERT_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER


def _add_row(ws, row, values, bold=False, subtotal=False):
    font = BOLD_FONT if bold or subtotal else NORMAL_FONT
    fill = GRIS_FILL if subtotal else None
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = font
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill
        if isinstance(val, (int, float)) and val != 0:
            cell.number_format = NUM_FMT
            cell.alignment = Alignment(horizontal='right')


def generer_boq_structure_xlsx(rs, params: dict, lang: str = "fr") -> bytes:
    """Generate BOQ Structure as Excel. Returns xlsx bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ Structure"

    d = rs.params
    boq = rs.boq
    poteaux = rs.poteaux
    poutre = rs.poutre_principale
    dalle = rs.dalle
    fond = rs.fondation

    # Pricing
    try:
        from prix_marche import get_prix_structure
        px = get_prix_structure(d.ville)
    except:
        class _PX:
            beton_c2530_m3=170000; beton_c3037_m3=185000; beton_c3545_m3=210000
            acier_ha400_kg=750; acier_ha500_kg=810; coffrage_bois_m2=18000
            coffrage_metal_m2=25000; terr_mecanique_m3=8500; terr_manuel_m3=5000
            remblai_m3=6500; pieu_fore_d600_ml=220000; pieu_fore_d800_ml=285000
            pieu_fore_d1000_ml=360000; semelle_filante_ml=85000; radier_m2=95000
            agglo_creux_10_m2=18000; agglo_creux_15_m2=24000; agglo_creux_20_m2=30000
            mo_chef_chantier_j=35000; mo_macon_j=18000; mo_ferrailleur_j=20000
            mo_manœuvre_j=8000
        px = _PX()

    prix_beton = {
        'C25/30': px.beton_c2530_m3, 'C30/37': px.beton_c3037_m3,
        'C35/45': px.beton_c3545_m3, 'C40/50': getattr(px, 'beton_c4050_m3', px.beton_c3545_m3),
    }.get(rs.classe_beton, px.beton_c3037_m3)
    prix_acier = px.acier_ha400_kg if rs.classe_acier == 'HA400' else px.acier_ha500_kg

    nb_pot = (d.nb_travees_x + 1) * (d.nb_travees_y + 1)
    surf_batie = boq.surface_batie_m2

    # Title rows
    ws.merge_cells('A1:G1')
    ws['A1'] = d.nom
    ws['A1'].font = Font(name="Calibri", size=14, bold=True)
    ws.merge_cells('A2:G2')
    ws['A2'] = f'{"BOQ Structure" if lang == "en" else "BOQ Structure"} — {d.ville} — R+{d.nb_niveaux-1} — {d.usage.value.capitalize()}'
    ws['A2'].font = Font(name="Calibri", size=11)
    ws.merge_cells('A3:G3')
    built_area_label = _t('Surface bâtie', 'Built area', lang)
    ws['A3'] = f'{built_area_label}: {surf_batie:,.0f} m² — {_t("Béton", "Concrete", lang)} {rs.classe_beton} — {_t("Acier", "Steel", lang)} {rs.classe_acier} — {_t("Prix 2026", "Pricing 2026", lang)}'
    ws['A3'].font = Font(name="Calibri", size=10, italic=True)

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 30

    if lang == 'en':
        headers = ['Lot', 'Description', 'Qty', 'Unit', 'Unit Price (FCFA)', 'Amount (FCFA)', 'Notes']
    else:
        headers = ['Lot', 'Désignation', 'Qté', 'Unité', 'P.U. (FCFA)', 'Montant (FCFA)', 'Observations']
    row = 5
    _add_header_row(ws, row, headers)
    row += 1

    grand_total = 0

    # LOT 1 — Installation
    inst_forfait = int(surf_batie * 2500)
    lot1_rows = [
        ('1.1', _t('Clôture de chantier', 'Site fence', lang), int(4*math.sqrt(d.surface_emprise_m2)), 'ml', 15000, int(4*math.sqrt(d.surface_emprise_m2)*15000), ''),
        ('1.2', _t('Base vie chantier', 'Temporary site facilities', lang), 1, _t('forfait', 'lump sum', lang), 0, int(surf_batie*800), _t('Modulaires', 'Modular units', lang)),
        ('1.3', _t('Branchements provisoires eau + élec', 'Temporary water + power connections', lang), 1, _t('forfait', 'lump sum', lang), 0, int(surf_batie*500), ''),
        ('1.4', _t('Signalétique sécurité + EPI', 'Safety signage + PPE', lang), 1, _t('forfait', 'lump sum', lang), 0, int(surf_batie*300), ''),
        ('1.5', _t('Repli et nettoyage', 'Demobilization & cleanup', lang), 1, _t('forfait', 'lump sum', lang), 0, int(surf_batie*600), ''),
    ]
    for vals in lot1_rows:
        _add_row(ws, row, vals)
        row += 1
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT 1', 'SUBTOTAL LOT 1', lang), '', '', '', inst_forfait, ''), subtotal=True)
    grand_total += inst_forfait
    row += 1

    # LOT 2 — Terrassement / Earthwork
    V_decap = d.surface_emprise_m2 * 0.30
    V_fouilles = d.surface_emprise_m2 * (fond.profondeur_m + 0.50)
    V_remblai = V_fouilles * 0.30
    V_evacu = V_fouilles * 0.70
    c_terr = int(V_decap*px.terr_mecanique_m3 + V_fouilles*px.terr_mecanique_m3 +
                 V_remblai*px.remblai_m3 + V_evacu*5000)
    lot2_rows = [
        ('2.1', _t('Décapage terre végétale e=30cm', 'Topsoil removal d=30cm', lang), int(V_decap), 'm³', px.terr_mecanique_m3, int(V_decap*px.terr_mecanique_m3), ''),
        ('2.2', _t('Fouilles générales mécaniques', 'Mechanical excavation', lang), int(V_fouilles), 'm³', px.terr_mecanique_m3, int(V_fouilles*px.terr_mecanique_m3), ''),
        ('2.3', _t('Remblai compacté', 'Compacted backfill', lang), int(V_remblai), 'm³', px.remblai_m3, int(V_remblai*px.remblai_m3), ''),
        ('2.4', _t('Évacuation terres', 'Soil evacuation', lang), int(V_evacu), 'm³', 5000, int(V_evacu*5000), ''),
    ]
    for vals in lot2_rows:
        _add_row(ws, row, vals)
        row += 1
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT 2', 'SUBTOTAL LOT 2', lang), '', '', '', c_terr, ''), subtotal=True)
    grand_total += c_terr
    row += 1

    # LOT 3 — Fondations / Foundations
    if fond.nb_pieux > 0:
        prix_pieu = {600: px.pieu_fore_d600_ml, 800: px.pieu_fore_d800_ml,
                     1000: px.pieu_fore_d1000_ml}.get(fond.diam_pieu_mm, px.pieu_fore_d800_ml)
        c_pieux = int(fond.nb_pieux * fond.longueur_pieu_m * prix_pieu)
        c_longr = int(nb_pot * 6 * 85000)
        c_fond = c_pieux + c_longr
        piles_label = _t('pieux', 'piles', lang)
        lot3_rows = [
            ('3.1', _t(f'Pieux forés Ø{fond.diam_pieu_mm}mm L={fond.longueur_pieu_m}m', f'Bored piles Ø{fond.diam_pieu_mm}mm L={fond.longueur_pieu_m}m', lang), int(fond.nb_pieux * fond.longueur_pieu_m), 'ml', prix_pieu, c_pieux, f'{fond.nb_pieux} {piles_label}'),
            ('3.2', _t('Longrines BA 30×50cm', 'Concrete pile caps 30×50cm', lang), int(nb_pot * 6), 'ml', 85000, c_longr, ''),
        ]
    else:
        c_fond = int(fond.beton_semelle_m3 * prix_beton * 1.6)
        lot3_rows = [
            ('3.1', _t('Béton de propreté e=10cm', 'Blinding concrete d=10cm', lang), int(d.surface_emprise_m2*0.10), 'm³', 120000, int(d.surface_emprise_m2*0.10*120000), ''),
            ('3.2', _t(f'Semelles BA {rs.classe_beton}', f'Concrete footings {rs.classe_beton}', lang), int(fond.beton_semelle_m3*0.6), 'm³', prix_beton, int(fond.beton_semelle_m3*0.6*prix_beton), ''),
            ('3.3', _t(f'Armatures semelles {rs.classe_acier}', f'Footing reinforcement {rs.classe_acier}', lang), int(fond.beton_semelle_m3*100), 'kg', prix_acier, int(fond.beton_semelle_m3*100*prix_acier), ''),
        ]
    for vals in lot3_rows:
        _add_row(ws, row, vals)
        row += 1
    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT 3', 'SUBTOTAL LOT 3', lang), '', '', '', c_fond, ''), subtotal=True)
    grand_total += c_fond
    row += 1

    # LOT 4 — Structure BA / Reinforced Concrete Structure
    ep_dalle_m = dalle.epaisseur_mm / 1000
    c_struct = 0

    # 4.1 — Poteaux béton + acier (per level — descente de charges)
    for i, pt in enumerate(poteaux):
        b = pt.section_mm / 1000
        V_niv = b**2 * d.hauteur_etage_m * nb_pot
        # Steel mass: bar area (cm²) → m² (/10000) × length × density × count
        As_niv = pt.nb_barres * math.pi * pt.diametre_mm**2 / 400 * nb_pot * d.hauteur_etage_m * 7850 / 10000
        c_b = int(V_niv * prix_beton)
        c_a = int(As_niv * prix_acier)
        cols_desc = _t(f'Poteaux béton {pt.section_mm}×{pt.section_mm} — {pt.niveau}',
                       f'Column concrete {pt.section_mm}×{pt.section_mm} — {pt.niveau}', lang)
        _add_row(ws, row, (f'4.1.{i*2+1}', cols_desc, f'{V_niv:.1f}', 'm³', prix_beton, c_b, ''))
        row += 1
        steel_desc = _t(f'Armatures poteaux {pt.niveau} — {pt.nb_barres}HA{pt.diametre_mm}',
                        f'Column reinforcement {pt.niveau} — {pt.nb_barres}HA{pt.diametre_mm}', lang)
        _add_row(ws, row, (f'4.1.{i*2+2}', steel_desc, f'{As_niv:.0f}', 'kg', prix_acier, c_a, f'{nb_pot} {_t("poteaux", "columns", lang)}'))
        row += 1
        c_struct += c_b + c_a

    # 4.2 — Poutres principales + secondaires (correct X/Y separation)
    pp_b = poutre.b_mm / 1000
    pp_h = poutre.h_mm / 1000
    px_m = getattr(d, 'portee_x_m', d.portee_max_m)
    py_m = getattr(d, 'portee_y_m', d.portee_max_m)
    # Poutres en X: (nb_travees_y + 1) rows × nb_travees_x spans × portee_x
    V_pp_x = pp_b * pp_h * px_m * (d.nb_travees_y + 1) * d.nb_travees_x * d.nb_niveaux
    # Poutres en Y: (nb_travees_x + 1) rows × nb_travees_y spans × portee_y
    V_pp_y = pp_b * pp_h * py_m * (d.nb_travees_x + 1) * d.nb_travees_y * d.nb_niveaux
    V_pp = V_pp_x + V_pp_y
    # Steel for beams: use As_inf + As_sup from calculation
    As_pp_cm2 = (poutre.As_inf_cm2 + poutre.As_sup_cm2) if hasattr(poutre, 'As_inf_cm2') else 10.0
    # Total beam length
    L_pp_total = (px_m * (d.nb_travees_y + 1) * d.nb_travees_x + py_m * (d.nb_travees_x + 1) * d.nb_travees_y) * d.nb_niveaux
    As_pp_kg = As_pp_cm2 / 10000 * L_pp_total * 7850
    c_pp_b = int(V_pp * prix_beton)
    c_pp_a = int(As_pp_kg * prix_acier)
    beams_desc = _t(f'Poutres principales béton {poutre.b_mm}×{poutre.h_mm}',
                    f'Main beams concrete {poutre.b_mm}×{poutre.h_mm}', lang)
    _add_row(ws, row, ('4.2.1', beams_desc, f'{V_pp:.1f}', 'm³', prix_beton, c_pp_b, ''))
    row += 1
    beams_steel = _t(f'Armatures poutres', f'Beam reinforcement', lang)
    _add_row(ws, row, ('4.2.2', beams_steel, f'{As_pp_kg:.0f}', 'kg', prix_acier, c_pp_a, f'As={As_pp_cm2:.1f}cm²'))
    row += 1
    c_struct += c_pp_b + c_pp_a

    # 4.3 — Dalle / Slab
    V_dalle = ep_dalle_m * surf_batie
    c_dalle = int(V_dalle * prix_beton)
    # Slab steel: As_x + As_y per m² × surface × density
    As_dalle_cm2 = (dalle.As_x_cm2_ml + dalle.As_y_cm2_ml) if hasattr(dalle, 'As_x_cm2_ml') else 8.0
    As_dalle_kg = As_dalle_cm2 / 10000 * surf_batie * 7850
    c_dalle_a = int(As_dalle_kg * prix_acier)
    slab_desc = _t(f'Dalle pleine ep.{dalle.epaisseur_mm}mm', f'Solid slab t={dalle.epaisseur_mm}mm', lang)
    _add_row(ws, row, ('4.3.1', slab_desc, f'{V_dalle:.1f}', 'm³', prix_beton, c_dalle, ''))
    row += 1
    slab_steel = _t('Armatures dalle (nappe inf. + sup.)', 'Slab reinforcement (top + bottom)', lang)
    _add_row(ws, row, ('4.3.2', slab_steel, f'{As_dalle_kg:.0f}', 'kg', prix_acier, c_dalle_a, f'As={As_dalle_cm2:.1f}cm²/ml'))
    row += 1
    c_struct += c_dalle + c_dalle_a

    # 4.4 — Coffrage / Formwork
    # Columns: perimeter × height × count × nb_niveaux
    coff_pot = 4 * (poteaux[0].section_mm / 1000) * d.hauteur_etage_m * nb_pot * d.nb_niveaux
    # Beams: (2 × h + b) × total beam length
    coff_pp = (2 * pp_h + pp_b) * L_pp_total
    # Slab: underside = surf_batie × nb_niveaux
    coff_dalle = surf_batie * d.nb_niveaux
    coff_total = coff_pot + coff_pp + coff_dalle
    c_coff = int(coff_total * px.coffrage_bois_m2)
    coff_desc = _t('Coffrage bois (poteaux + poutres + dalle)', 'Formwork (columns + beams + slab)', lang)
    _add_row(ws, row, ('4.4', coff_desc, f'{coff_total:.0f}', 'm²', px.coffrage_bois_m2, c_coff,
                       _t(f'Pot:{coff_pot:.0f} + Poutr:{coff_pp:.0f} + Dalle:{coff_dalle:.0f}',
                          f'Col:{coff_pot:.0f} + Beam:{coff_pp:.0f} + Slab:{coff_dalle:.0f}', lang)))
    row += 1
    c_struct += c_coff

    _add_row(ws, row, ('', _t('SOUS-TOTAL LOT 4', 'SUBTOTAL LOT 4', lang), '', '', '', c_struct, ''), subtotal=True)
    grand_total += c_struct
    row += 2

    # GRAND TOTAL / GRAND TOTAL
    total_label = _t('TOTAL GÉNÉRAL STRUCTURE', 'TOTAL STRUCTURAL COST', lang)
    unit_label = _t('FCFA/m²', 'FCFA/m²', lang)
    _add_row(ws, row, ('', total_label, '', '', '', grand_total, f'{grand_total/surf_batie:,.0f} {unit_label}'), bold=True)
    ws.cell(row=row, column=6).font = Font(name="Calibri", size=12, bold=True, color="43A956")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
